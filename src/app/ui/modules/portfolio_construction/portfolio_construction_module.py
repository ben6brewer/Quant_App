"""Portfolio Construction Module - Main Orchestrator"""

import csv
from datetime import datetime
from typing import List, Dict, Any, Tuple

from PySide6.QtWidgets import QWidget, QVBoxLayout, QStackedWidget, QApplication, QFileDialog
from PySide6.QtCore import Qt

from app.core.theme_manager import ThemeManager
from app.ui.widgets.common import CustomMessageBox
from app.ui.widgets.common.loading_overlay import LoadingOverlay
from app.ui.widgets.common.lazy_theme_mixin import LazyThemeMixin

from .services import PortfolioService, PortfolioPersistence, PortfolioSettingsManager
from .widgets import (
    PortfolioControls,
    TransactionLogTable,
    AggregatePortfolioTable,
    NewPortfolioDialog,
    LoadPortfolioDialog,
    RenamePortfolioDialog,
    ImportPortfolioDialog,
    ExportDialog,
    ViewTabBar
)
from .widgets.portfolio_settings_dialog import PortfolioSettingsDialog


class PortfolioConstructionModule(LazyThemeMixin, QWidget):
    """
    Main portfolio construction module.
    Orchestrates all widgets and services for portfolio management.
    """

    def __init__(self, theme_manager: ThemeManager, parent=None):
        super().__init__(parent)
        self.theme_manager = theme_manager
        self._theme_dirty = False  # For lazy theme application

        # Initialize services
        PortfolioPersistence.initialize()

        # State
        self.current_portfolio = None  # Current portfolio dict
        self.unsaved_changes = False

        # Price cache - only fetch when tickers change
        self._cached_prices = {}  # ticker -> price
        self._cached_tickers = set()  # Set of tickers we've fetched prices for

        # Name cache - ticker short names from Yahoo Finance
        self._cached_names = {}  # ticker -> short name

        # Settings manager (handles persistence)
        self._settings_manager = PortfolioSettingsManager()

        # Loading overlay (created on demand)
        self._loading_overlay = None

        self._setup_ui()

        # Apply persisted settings to widgets
        self._apply_settings()
        self._connect_signals()
        self._apply_theme()

        # Initialize portfolio list without loading data
        self._initialize_portfolio_list()

        # Set initial view mode (Transaction Log by default)
        self.controls.set_view_mode(is_transaction_view=True)

        # Connect theme changes (lazy - only apply when visible)
        self.theme_manager.theme_changed.connect(self._on_theme_changed_lazy)

    def showEvent(self, event):
        """Handle show event - apply pending theme if needed."""
        super().showEvent(event)
        self._check_theme_dirty()

    def _setup_ui(self):
        """Setup main UI layout."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Controls bar at top
        self.controls = PortfolioControls(self.theme_manager)
        layout.addWidget(self.controls)

        # View tab bar
        self.view_tab_bar = ViewTabBar(self.theme_manager)
        layout.addWidget(self.view_tab_bar)

        # Stacked widget for switching between tables
        self.table_stack = QStackedWidget()

        # Index 0: Transaction Log (full screen, no label)
        self.transaction_table = TransactionLogTable(self.theme_manager)
        self.table_stack.addWidget(self.transaction_table)

        # Index 1: Portfolio Holdings (full screen, no label)
        self.aggregate_table = AggregatePortfolioTable(self.theme_manager)
        self.table_stack.addWidget(self.aggregate_table)

        # Default to Transaction Log
        self.table_stack.setCurrentIndex(0)

        layout.addWidget(self.table_stack)

    def _connect_signals(self):
        """Connect all signals."""
        # Controls
        self.controls.portfolio_changed.connect(self._on_portfolio_changed)
        self.controls.save_clicked.connect(self._save_portfolio)
        self.controls.import_clicked.connect(self._import_portfolio_dialog)
        self.controls.export_clicked.connect(self._export_dialog)
        self.controls.new_portfolio_clicked.connect(self._new_portfolio_dialog)
        self.controls.rename_portfolio_clicked.connect(self._rename_portfolio_dialog)
        self.controls.delete_portfolio_clicked.connect(self._delete_portfolio_dialog)
        self.controls.home_clicked.connect(self._on_home_clicked)
        self.controls.settings_clicked.connect(self._open_settings_dialog)

        # Transaction table
        self.transaction_table.transaction_added.connect(self._on_transaction_changed)
        self.transaction_table.transaction_modified.connect(self._on_transaction_changed)
        self.transaction_table.transaction_deleted.connect(self._on_transaction_changed)

        # View tab bar
        self.view_tab_bar.view_changed.connect(self._on_view_changed)

    def _on_view_changed(self, index: int):
        """Handle view tab change."""
        self.table_stack.setCurrentIndex(index)
        # Show editing buttons only on Transaction Log view (index 0)
        self.controls.set_view_mode(is_transaction_view=(index == 0))

    def _initialize_portfolio_list(self):
        """Initialize portfolio dropdown without loading any portfolio."""
        portfolios = PortfolioPersistence.list_portfolios_by_recent()

        # Populate dropdown with no portfolio selected (shows placeholder)
        self.controls.update_portfolio_list(portfolios, None)

        # Show empty state - user must explicitly select a portfolio
        self._show_empty_state()

    def _populate_transaction_table(self):
        """Populate transaction table from current portfolio."""
        self.transaction_table.clear_all_transactions()

        if not self.current_portfolio:
            return

        transactions = self.current_portfolio.get("transactions", [])

        # Initialize sequence counter from existing transactions
        self.transaction_table._initialize_sequence_counter(transactions)

        # Use batch loading mode to avoid O(N²) FREE CASH updates
        self.transaction_table.begin_batch_loading()
        for transaction in transactions:
            self.transaction_table.add_transaction_row(transaction)
        self.transaction_table.end_batch_loading()

        # Ensure blank row exists for adding new transactions
        self.transaction_table._ensure_blank_row()

        # Sort by date descending (most recent first) with sequence for same-day ordering
        self.transaction_table.sort_by_date_descending()

        # Historical prices fetched in _update_aggregate_table() to avoid duplicate calls
        self.unsaved_changes = False

    def _on_transaction_changed(self, *args):
        """Handle transaction add/modify/delete."""
        self.unsaved_changes = True
        self._update_aggregate_table()

    def _show_empty_state(self):
        """Display empty state when no portfolio loaded."""
        self.current_portfolio = None
        self.transaction_table.clear_all_transactions()
        self.aggregate_table.setRowCount(0)
        self.unsaved_changes = False
        # Clear price cache
        self._cached_prices.clear()
        self._cached_tickers.clear()
        # Update button states (disable Save/Rename/Delete)
        self.controls._update_button_states(False)

    def _update_aggregate_table(self, force_fetch: bool = False):
        """
        Recalculate and update aggregate table.

        Args:
            force_fetch: If True, fetch prices for all tickers (used on portfolio load)
        """
        transactions = self.transaction_table.get_all_transactions()

        if not transactions:
            self.aggregate_table.setRowCount(0)
            return

        # Get unique tickers (excluding FREE CASH - it doesn't need price fetching)
        tickers = set(
            t["ticker"] for t in transactions
            if t["ticker"] and t["ticker"].upper() != PortfolioService.FREE_CASH_TICKER
        )

        # Determine which tickers need price fetching
        if tickers:
            if force_fetch:
                # Full refresh - fetch all tickers
                tickers_to_fetch = list(tickers)
            else:
                # Only fetch prices for NEW tickers (not already cached)
                tickers_to_fetch = [t for t in tickers if t not in self._cached_tickers]

            # Fetch prices and names only for tickers that need it
            if tickers_to_fetch:
                new_prices = PortfolioService.fetch_current_prices(tickers_to_fetch)
                new_names = PortfolioService.fetch_ticker_names(tickers_to_fetch)
                # Update caches
                self._cached_prices.update(new_prices)
                self._cached_names.update(new_names)
                self._cached_tickers.update(tickers_to_fetch)

            # Remove cached tickers that are no longer in use
            removed_tickers = self._cached_tickers - tickers
            for ticker in removed_tickers:
                self._cached_tickers.discard(ticker)
                self._cached_prices.pop(ticker, None)
                self._cached_names.pop(ticker, None)

        # Use cached prices for calculations
        current_prices = {t: self._cached_prices.get(t) for t in tickers}

        # Use cached names for display
        ticker_names = {t: self._cached_names.get(t) for t in tickers}

        # Update transaction table with current prices and names
        self.transaction_table.update_current_prices(current_prices)
        self.transaction_table.update_ticker_names(ticker_names)

        # Also fetch historical prices for new ticker/date combinations
        # (batch fetch handles caching internally)
        self.transaction_table.fetch_historical_prices_batch()

        # Calculate holdings (excludes FREE CASH)
        holdings = PortfolioService.calculate_aggregate_holdings(transactions, current_prices)

        # Calculate FREE CASH summary
        free_cash_summary = PortfolioService.calculate_free_cash_summary(transactions)

        # Update aggregate table with holdings, FREE CASH, and ticker names
        self.aggregate_table.update_holdings(holdings, free_cash_summary, ticker_names)

    def _refresh_prices(self):
        """Manually refresh current prices and names."""
        # Clear caches and force fetch all
        self._cached_prices.clear()
        self._cached_names.clear()
        self._cached_tickers.clear()
        self._update_aggregate_table(force_fetch=True)
        CustomMessageBox.information(
            self.theme_manager,
            self,
            "Prices Refreshed",
            "Current prices updated successfully."
        )

    def _save_portfolio(self):
        """Save current portfolio to disk."""
        if not self.current_portfolio:
            return

        # Update transactions in portfolio
        self.current_portfolio["transactions"] = self.transaction_table.get_all_transactions()

        # Save
        success = PortfolioPersistence.save_portfolio(self.current_portfolio)

        if success:
            self.unsaved_changes = False
            CustomMessageBox.information(
                self.theme_manager,
                self,
                "Saved",
                f"Portfolio '{self.current_portfolio['name']}' saved successfully."
            )
        else:
            CustomMessageBox.critical(
                self.theme_manager,
                self,
                "Save Error",
                "Failed to save portfolio."
            )

    def _load_portfolio_dialog(self):
        """Open load portfolio dialog."""
        # Check for unsaved changes
        if self.unsaved_changes:
            reply = CustomMessageBox.question(
                self.theme_manager,
                self,
                "Unsaved Changes",
                "You have unsaved changes. Do you want to save before loading?",
                CustomMessageBox.Yes | CustomMessageBox.No | CustomMessageBox.Cancel,
                CustomMessageBox.Cancel
            )

            if reply == CustomMessageBox.Yes:
                self._save_portfolio()
            elif reply == CustomMessageBox.Cancel:
                return

        # Open dialog
        portfolios = PortfolioPersistence.list_portfolios_by_recent()
        if not portfolios:
            CustomMessageBox.information(
                self.theme_manager,
                self,
                "No Portfolios",
                "No portfolios found. Create a new one first."
            )
            return

        dialog = LoadPortfolioDialog(self.theme_manager, portfolios, self)

        if dialog.exec():
            name = dialog.get_selected_name()
            if name:
                self._load_portfolio(name)

    def _load_portfolio(self, name: str):
        """Load portfolio by name."""
        portfolio = PortfolioPersistence.load_portfolio(name)

        if not portfolio:
            CustomMessageBox.critical(
                self.theme_manager,
                self,
                "Load Error",
                f"Failed to load portfolio '{name}'."
            )
            return

        # Show loading overlay
        self._show_loading_overlay()

        try:
            # Clear caches when loading new portfolio
            self._cached_prices.clear()
            self._cached_names.clear()
            self._cached_tickers.clear()

            self.current_portfolio = portfolio
            self._populate_transaction_table()
            # Force fetch all prices on portfolio load
            self._update_aggregate_table(force_fetch=True)

            # Record visit for recent ordering
            PortfolioPersistence.record_visit(name)

            # Update controls and enable buttons (with recent ordering)
            portfolios = PortfolioPersistence.list_portfolios_by_recent()
            self.controls.update_portfolio_list(portfolios, name)
            self.controls._update_button_states(True)
        finally:
            # Hide loading overlay
            self._hide_loading_overlay()

    def _show_loading_overlay(self, message: str = "Loading Portfolio..."):
        """Show loading overlay over the tab bar and table area."""
        if self._loading_overlay is None:
            # Parent to self (the module) to cover both tab bar and tables
            self._loading_overlay = LoadingOverlay(
                self, self.theme_manager, message
            )

        # Calculate rect that covers view_tab_bar + table_stack
        # Get positions relative to self
        tab_bar_top = self.view_tab_bar.geometry().top()
        table_bottom = self.table_stack.geometry().bottom()
        content_rect = self.rect()
        content_rect.setTop(tab_bar_top)
        content_rect.setBottom(table_bottom)

        self._loading_overlay.setGeometry(content_rect)
        self._loading_overlay.show()
        self._loading_overlay.raise_()
        # Force UI update to render overlay before heavy work
        QApplication.processEvents()

    def _hide_loading_overlay(self):
        """Hide and cleanup loading overlay."""
        if self._loading_overlay is not None:
            self._loading_overlay.hide()
            self._loading_overlay.deleteLater()
            self._loading_overlay = None

    def _new_portfolio_dialog(self):
        """Open new portfolio dialog."""
        # Check for unsaved changes
        if self.unsaved_changes:
            reply = CustomMessageBox.question(
                self.theme_manager,
                self,
                "Unsaved Changes",
                "You have unsaved changes. Do you want to save before creating a new portfolio?",
                CustomMessageBox.Yes | CustomMessageBox.No | CustomMessageBox.Cancel,
                CustomMessageBox.Cancel
            )

            if reply == CustomMessageBox.Yes:
                self._save_portfolio()
            elif reply == CustomMessageBox.Cancel:
                return

        # Open dialog
        existing = PortfolioPersistence.list_portfolios()
        dialog = NewPortfolioDialog(self.theme_manager, existing, self)

        if dialog.exec():
            name = dialog.get_name()
            # Create new portfolio
            self.current_portfolio = PortfolioPersistence.create_new_portfolio(name)
            PortfolioPersistence.save_portfolio(self.current_portfolio)

            # Clear tables and caches
            self.transaction_table.clear_all_transactions()
            self.aggregate_table.setRowCount(0)
            self._cached_prices.clear()
            self._cached_names.clear()
            self._cached_tickers.clear()

            # Ensure blank row exists for immediate editing
            self.transaction_table._ensure_blank_row()

            # Record visit for recent ordering
            PortfolioPersistence.record_visit(name)

            # Update controls and enable buttons (with recent ordering)
            portfolios = PortfolioPersistence.list_portfolios_by_recent()
            self.controls.update_portfolio_list(portfolios, name)
            self.controls._update_button_states(True)

            self.unsaved_changes = False

    def _rename_portfolio_dialog(self):
        """Open rename portfolio dialog."""
        if not self.current_portfolio:
            return

        current_name = self.current_portfolio.get("name", "")
        existing = PortfolioPersistence.list_portfolios()
        dialog = RenamePortfolioDialog(self.theme_manager, current_name, existing, self)

        if dialog.exec():
            new_name = dialog.get_name()
            if new_name and new_name != current_name:
                success = PortfolioPersistence.rename_portfolio(current_name, new_name)
                if success:
                    # Update current portfolio name
                    self.current_portfolio["name"] = new_name
                    # Refresh dropdown (with recent ordering)
                    portfolios = PortfolioPersistence.list_portfolios_by_recent()
                    self.controls.update_portfolio_list(portfolios, new_name)
                else:
                    CustomMessageBox.critical(
                        self.theme_manager,
                        self,
                        "Rename Error",
                        f"Failed to rename portfolio to '{new_name}'."
                    )

    def _delete_portfolio_dialog(self):
        """Show confirmation dialog and delete current portfolio."""
        if not self.current_portfolio:
            return

        portfolio_name = self.current_portfolio.get("name", "")

        # Confirm deletion
        reply = CustomMessageBox.question(
            self.theme_manager,
            self,
            "Delete Portfolio",
            f"Are you sure you want to delete '{portfolio_name}'?\n\nThis action cannot be undone.",
            CustomMessageBox.Yes | CustomMessageBox.No,
            CustomMessageBox.No
        )

        if reply != CustomMessageBox.Yes:
            return

        # Delete portfolio
        success = PortfolioPersistence.delete_portfolio(portfolio_name)

        if success:
            # Remove from recent visits
            PortfolioPersistence.remove_from_recent(portfolio_name)

            # Get remaining portfolios (with recent ordering)
            portfolios = PortfolioPersistence.list_portfolios_by_recent()

            # Update dropdown (no portfolio selected)
            self.controls.update_portfolio_list(portfolios, None)

            # Show empty state - user must select another portfolio
            self._show_empty_state()
        else:
            CustomMessageBox.critical(
                self.theme_manager,
                self,
                "Delete Error",
                f"Failed to delete portfolio '{portfolio_name}'."
            )

    def _import_portfolio_dialog(self):
        """Show import dialog and process import."""
        # Validate portfolio is loaded
        if not self.current_portfolio:
            CustomMessageBox.warning(
                self.theme_manager,
                self,
                "No Portfolio",
                "Please load or create a portfolio first."
            )
            return

        # Get available portfolios (exclude current)
        current_name = self.current_portfolio.get("name", "")
        all_portfolios = PortfolioPersistence.list_portfolios()
        available = [p for p in all_portfolios if p != current_name]

        if not available:
            CustomMessageBox.warning(
                self.theme_manager,
                self,
                "No Portfolios",
                "No other portfolios available to import from."
            )
            return

        # Show dialog
        dialog = ImportPortfolioDialog(self.theme_manager, available, self)
        if dialog.exec() != 1:  # QDialog.Accepted = 1
            return

        config = dialog.get_import_config()
        if not config:
            return

        # Load source transactions
        source_data = PortfolioPersistence.load_portfolio(config["source_portfolio"])
        if not source_data:
            CustomMessageBox.critical(
                self.theme_manager,
                self,
                "Load Error",
                f"Failed to load source portfolio '{config['source_portfolio']}'."
            )
            return

        source_txs = source_data.get("transactions", [])

        if not source_txs:
            CustomMessageBox.information(
                self.theme_manager,
                self,
                "Empty Portfolio",
                "Source portfolio has no transactions to import."
            )
            return

        # Process based on mode
        if config["import_mode"] == "flat":
            new_txs = PortfolioService.process_flat_import(
                source_txs,
                config["include_fees"],
                config["skip_zero_positions"]
            )
        else:
            new_txs = PortfolioService.generate_imported_transactions(
                source_txs,
                config["include_fees"]
            )

        if not new_txs:
            CustomMessageBox.information(
                self.theme_manager,
                self,
                "No Transactions",
                "No transactions to import after processing."
            )
            return

        # Show loading overlay during import
        self._show_loading_overlay("Importing Transactions...")

        try:
            # Add to current portfolio using batch mode for O(N) performance
            self.transaction_table.begin_batch_loading()
            for tx in new_txs:
                self.transaction_table.add_transaction_row(tx)
            self.transaction_table.end_batch_loading()

            # Sort to maintain date-descending order after import
            self.transaction_table.sort_by_date_descending()

            self.unsaved_changes = True

            # Update aggregate table
            self._update_aggregate_table()

            # Fetch historical prices for new transactions
            self.transaction_table.fetch_historical_prices_batch()
        finally:
            # Hide loading overlay
            self._hide_loading_overlay()

        # Show success message
        count = len(new_txs)
        mode_desc = "consolidated positions" if config["import_mode"] == "flat" else "transactions"
        CustomMessageBox.information(
            self.theme_manager,
            self,
            "Import Complete",
            f"Successfully imported {count} {mode_desc} from '{config['source_portfolio']}'."
        )

    def _on_portfolio_changed(self, name: str):
        """Handle portfolio selection change in dropdown."""
        # Strip "[Port] " prefix if present
        if name.startswith("[Port] "):
            name = name[7:]

        if not name:
            return

        # Track if this is first load
        is_first_load = self.current_portfolio is None

        # Skip same-portfolio check on first load
        if not is_first_load and self.current_portfolio and name == self.current_portfolio.get("name"):
            return

        # Check for unsaved changes (skip on first load)
        if not is_first_load and self.unsaved_changes:
            reply = CustomMessageBox.question(
                self.theme_manager,
                self,
                "Unsaved Changes",
                "You have unsaved changes. Do you want to save before switching?",
                CustomMessageBox.Yes | CustomMessageBox.No | CustomMessageBox.Cancel,
                CustomMessageBox.Cancel
            )

            if reply == CustomMessageBox.Yes:
                self._save_portfolio()
            elif reply == CustomMessageBox.Cancel:
                # Revert dropdown (with recent ordering)
                current_name = self.current_portfolio.get("name") if self.current_portfolio else None
                self.controls.update_portfolio_list(
                    PortfolioPersistence.list_portfolios_by_recent(),
                    current_name
                )
                return

        self._load_portfolio(name)

    def _on_home_clicked(self):
        """Handle home button click."""
        parent = self.parent()
        while parent:
            if hasattr(parent, 'show_home'):
                parent.show_home()
                break
            parent = parent.parent()

    def _apply_theme(self):
        """Apply theme to module."""
        theme = self.theme_manager.current_theme

        if theme == "light":
            bg_color = "#ffffff"
        elif theme == "bloomberg":
            bg_color = "#000814"
        else:
            bg_color = "#1e1e1e"

        self.setStyleSheet(f"""
            QWidget {{
                background-color: {bg_color};
            }}
        """)

    def _open_settings_dialog(self):
        """Open settings dialog."""
        dialog = PortfolioSettingsDialog(
            self.theme_manager,
            self._settings_manager.get_all_settings(),
            self
        )

        if dialog.exec():
            new_settings = dialog.get_settings()
            if new_settings:
                # Save settings to disk
                self._settings_manager.update_settings(new_settings)
                # Apply settings to widgets
                self._apply_settings()

    def _apply_settings(self):
        """Apply current settings to widgets."""
        self.transaction_table.set_highlight_editable(
            self._settings_manager.get_setting("highlight_editable_fields")
        )
        self.transaction_table.set_hide_free_cash_summary(
            self._settings_manager.get_setting("hide_free_cash_summary")
        )

    # ========== Export Methods ==========

    def _export_dialog(self):
        """Show export format dialog and process export."""
        if not self.current_portfolio:
            CustomMessageBox.warning(
                self.theme_manager,
                self,
                "No Portfolio",
                "Please load or create a portfolio first."
            )
            return

        dialog = ExportDialog(self.theme_manager, self)
        if dialog.exec() != 1:  # QDialog.Accepted
            return

        export_format = dialog.get_format()
        if not export_format:
            return

        # Get data based on current view
        current_view = self.table_stack.currentIndex()

        if current_view == 0:  # Transaction Log
            data, columns, prefix = self._get_transaction_export_data()
        else:  # Holdings
            data, columns, prefix = self._get_holdings_export_data()

        if not data:
            CustomMessageBox.information(
                self.theme_manager,
                self,
                "No Data",
                "No data to export."
            )
            return

        if export_format == "csv":
            self._export_to_csv(data, columns, prefix)
        else:
            self._export_to_excel(data, columns, prefix)

    def _get_transaction_export_data(self) -> Tuple[List[Dict[str, Any]], List[str], str]:
        """
        Get transaction data for export.

        Returns:
            Tuple of (data rows, column headers, filename prefix)
        """
        transactions = self.transaction_table.get_all_transactions()

        columns = [
            "Date", "Ticker", "Name", "Quantity", "Execution Price",
            "Fees", "Type", "Daily Closing Price", "Live Price",
            "Principal", "Market Value"
        ]

        export_data = []
        for tx in transactions:
            ticker = tx.get("ticker", "")
            quantity = tx.get("quantity", 0) or 0
            entry_price = tx.get("entry_price", 0) or 0
            fees = tx.get("fees", 0) or 0
            current_price = self._cached_prices.get(ticker, 0) or 0

            # Get historical price for this transaction date
            tx_date = tx.get("date", "")
            historical_prices = self.transaction_table._historical_prices
            daily_close = historical_prices.get(ticker, {}).get(tx_date, "")

            # Calculate principal and market value
            principal = quantity * entry_price + fees
            market_value = quantity * current_price if current_price else ""

            row = {
                "Date": tx_date,
                "Ticker": ticker,
                "Name": self._cached_names.get(ticker, ""),
                "Quantity": quantity,
                "Execution Price": entry_price,
                "Fees": fees,
                "Type": tx.get("transaction_type", ""),
                "Daily Closing Price": daily_close,
                "Live Price": current_price if current_price else "",
                "Principal": principal,
                "Market Value": market_value
            }
            export_data.append(row)

        portfolio_name = self.current_portfolio.get("name", "portfolio")
        return export_data, columns, f"{portfolio_name}_transactions"

    def _get_holdings_export_data(self) -> Tuple[List[Dict[str, Any]], List[str], str]:
        """
        Get holdings data for export (excludes TOTAL row).

        Returns:
            Tuple of (data rows, column headers, filename prefix)
        """
        holdings = self.aggregate_table._holdings_data

        columns = [
            "Ticker", "Name", "Quantity", "Avg Cost Basis",
            "Current Price", "Market Value", "P&L", "Weight %"
        ]

        export_data = []
        for holding in holdings:
            ticker = holding.get("ticker", "")
            is_free_cash = holding.get("_is_free_cash", False)

            row = {
                "Ticker": ticker,
                "Name": self._cached_names.get(ticker, "") if not is_free_cash else "",
                "Quantity": holding.get("total_quantity", 0),
                "Avg Cost Basis": holding.get("avg_cost_basis", 0),
                "Current Price": holding.get("current_price", "") if not is_free_cash else "",
                "Market Value": holding.get("market_value", ""),
                "P&L": holding.get("total_pnl", 0) if not is_free_cash else "",
                "Weight %": holding.get("weight_pct", 0)
            }
            export_data.append(row)

        portfolio_name = self.current_portfolio.get("name", "portfolio")
        return export_data, columns, f"{portfolio_name}_holdings"

    def _export_to_csv(self, data: List[Dict[str, Any]], columns: List[str], filename_prefix: str):
        """
        Export data to CSV file.

        Args:
            data: List of row dictionaries
            columns: List of column headers
            filename_prefix: Default filename without extension
        """
        default_name = f"{filename_prefix}.csv"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to CSV",
            default_name,
            "CSV Files (*.csv);;All Files (*)"
        )

        if not file_path:
            return  # User cancelled

        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                writer.writeheader()
                writer.writerows(data)

            CustomMessageBox.information(
                self.theme_manager,
                self,
                "Export Complete",
                f"Data exported to:\n{file_path}"
            )
        except Exception as e:
            CustomMessageBox.critical(
                self.theme_manager,
                self,
                "Export Error",
                f"Failed to export data:\n{str(e)}"
            )

    def _export_to_excel(self, data: List[Dict[str, Any]], columns: List[str], sheet_name: str):
        """
        Export data to Excel.

        On Windows: Uses COM automation to open Excel directly.
        On macOS/Linux: Creates .xlsx file with openpyxl and opens with default app.

        Args:
            data: List of row dictionaries
            columns: List of column headers
            sheet_name: Name for the worksheet
        """
        import sys

        if sys.platform == 'win32':
            self._export_to_excel_windows(data, columns, sheet_name)
        else:
            self._export_to_excel_crossplatform(data, columns, sheet_name)

    def _export_to_excel_windows(self, data: List[Dict[str, Any]], columns: List[str], sheet_name: str):
        """Export to Excel using Windows COM automation."""
        try:
            import win32com.client
        except ImportError:
            # Fall back to cross-platform method
            self._export_to_excel_crossplatform(data, columns, sheet_name)
            return

        try:
            # Create Excel application
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = True

            # Create new workbook
            workbook = excel.Workbooks.Add()
            worksheet = workbook.ActiveSheet
            worksheet.Name = sheet_name[:31]  # Excel sheet names max 31 chars

            # Write headers
            for col_idx, col_name in enumerate(columns, start=1):
                worksheet.Cells(1, col_idx).Value = col_name
                worksheet.Cells(1, col_idx).Font.Bold = True

            # Write data
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    value = row_data.get(col_name, "")
                    worksheet.Cells(row_idx, col_idx).Value = value

            # Auto-fit columns
            worksheet.Columns.AutoFit()

        except Exception as e:
            CustomMessageBox.critical(
                self.theme_manager,
                self,
                "Excel Error",
                f"Failed to open Excel:\n{str(e)}\n\n"
                "Make sure Microsoft Excel is installed."
            )

    def _export_to_excel_crossplatform(self, data: List[Dict[str, Any]], columns: List[str], sheet_name: str):
        """Export to Excel using openpyxl (works on macOS/Linux)."""
        import subprocess
        import sys
        import tempfile
        import os

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            CustomMessageBox.warning(
                self.theme_manager,
                self,
                "Excel Not Available",
                "Excel export requires the 'openpyxl' package.\n\n"
                "Install it with: pip install openpyxl"
            )
            return

        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]  # Excel sheet names max 31 chars

            # Write headers with bold font
            bold_font = Font(bold=True)
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = bold_font

            # Write data
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    value = row_data.get(col_name, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-fit columns (approximate)
            for col_idx, col_name in enumerate(columns, start=1):
                max_length = len(str(col_name))
                for row_data in data:
                    cell_value = str(row_data.get(col_name, ""))
                    max_length = max(max_length, len(cell_value))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

            # Save to temp file
            temp_dir = tempfile.gettempdir()
            file_path = os.path.join(temp_dir, f"{sheet_name}.xlsx")
            wb.save(file_path)

            # Open with default application
            if sys.platform == 'darwin':  # macOS
                subprocess.run(['open', file_path], check=True)
            elif sys.platform == 'win32':  # Windows fallback
                os.startfile(file_path)
            else:  # Linux
                subprocess.run(['xdg-open', file_path], check=True)

        except Exception as e:
            CustomMessageBox.critical(
                self.theme_manager,
                self,
                "Excel Error",
                f"Failed to create Excel file:\n{str(e)}"
            )
